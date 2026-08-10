import json, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

R = "/Users/Shared/ClaudeShared/par-build"
OUT = ("/Users/stowaway/Library/Application Support/Claude/local-agent-mode-sessions/"
       "d8234097-867b-450a-a94a-7039b7206977/41d7a41d-27ce-46ff-980a-de9213a09d79/"
       "local_7312f944-801e-43f1-a66e-ec349d2ed4e5/outputs/Par_Levels_2026-08-09.xlsx")

FONT = "Arial"
HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=10)
UP_FILL = PatternFill("solid", fgColor="E2EFDA")
DN_FILL = PatternFill("solid", fgColor="FCE4D6")
OV_FILL = PatternFill("solid", fgColor="FFF2CC")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COLS = [
    ("SKU", 42), ("Reporting group", 20), ("Live par", 10), ("Recommended", 13),
    ("Delta", 10), ("Change %", 10), ("Demand /wk", 11), ("Shrink /wk", 11),
    ("Seasonal", 10), ("Method", 15), ("Service", 10), ("Burst floor", 11),
    ("Override", 12), ("Flags", 34),
]


def sheet_for(wb, venue_file, title):
    d = json.load(open(f"{R}/data/par_recommendations_{venue_file}.json"))
    skus = d["skus"]

    # Decisions at the top: real changes to live pars first (biggest move first),
    # then unchanged, then SKUs with no live par at all (noise, not decisions).
    def sort_key(r):
        c = r.get("current_par")
        if c is None:
            return (2, -r["rec_par"], r["product"])
        dd = abs(r["rec_par"] - c)
        if dd < 1e-9:
            return (1, 0.0, r["product"])
        return (0, -dd, r["product"])
    skus = sorted(skus, key=sort_key)

    ws = wb.create_sheet(title)
    ws.append([c[0] for c in COLS])
    for i, (name, w) in enumerate(COLS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
        cell = ws.cell(row=1, column=i)
        cell.fill = HDR_FILL; cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"

    for r in skus:
        sv = r.get("service") or {}
        sh = r.get("shrinkage") or {}
        ov = r.get("override")
        ws.append([
            r["product"], r.get("reporting_group") or "", r.get("current_par"), r["rec_par"],
            None, None,
            round((r.get("drivers") or {}).get("true_wk") or 0, 2),
            round(sh.get("loss_per_week") or 0, 3),
            r.get("seasonal_index"),
            (r.get("forecast_method") or "").replace("Deseasonalised level x seasonal index", "Seasonal"),
            sv.get("service_class") or "", sv.get("burst_floor") or 0,
            f"{ov['type']} {ov['value']}" if ov else "",
            ", ".join(r.get("flags") or []),
        ])
        i = ws.max_row
        # Values, not formulas: no LibreOffice on this host to recalculate, and an
        # uncalculated formula reads back blank in previews. The model is the
        # source of truth for these numbers, not the sheet.
        cur_v = r.get("current_par")
        if cur_v is None:
            ws[f"E{i}"] = "new"; ws[f"F{i}"] = ""
        else:
            ws[f"E{i}"] = round(r["rec_par"] - cur_v, 1)
            ws[f"F{i}"] = (r["rec_par"] / cur_v - 1) if cur_v > 0 else ""
        for col in range(1, len(COLS) + 1):
            c = ws.cell(row=i, column=col)
            c.font = Font(name=FONT, size=10); c.border = BORDER
        for col in ("C", "D", "E", "G", "H", "L"):
            ws[f"{col}{i}"].number_format = "0.0"
        ws[f"F{i}"].number_format = "0%"
        ws[f"I{i}"].number_format = "0.00"
        if cur_v is not None:
            if r["rec_par"] > cur_v + 1e-9:
                ws[f"D{i}"].fill = UP_FILL
            elif r["rec_par"] < cur_v - 1e-9:
                ws[f"D{i}"].fill = DN_FILL
        if ov:
            ws[f"M{i}"].fill = OV_FILL
    ws.auto_filter.ref = f"A1:N{ws.max_row}"
    return d


wb = Workbook()
wb.remove(wb.active)
stow = sheet_for(wb, "stowaway", "Stowaway")
hg = sheet_for(wb, "harry_gatos", "Harry Gatos")

ws = wb.create_sheet("Read me", 0)
ws.column_dimensions["A"].width = 34
ws.column_dimensions["B"].width = 66
exp = stow["exposure"]
rows = [
    ("Par levels — recommendation", ""),
    ("Generated", stow["generated_at"]),
    ("Engine", stow.get("engine", "v3")),
    ("Sales data", f"{stow['weeks']} weeks, {stow['week_range'][0]} to {stow['week_range'][1]}"),
    ("Order cycle", f"order Sun {stow['order_sunday']} -> {exp['day_units']} weighted day-units "
                    f"({exp['day_units']/exp['normal_day_units']:.2f}x a normal cycle)"),
    ("", ""),
    ("STOWAWAY", f"{stow['summary']['n_skus']} SKUs — {stow['summary']['increase']} up, "
                 f"{stow['summary']['decrease']} down, {stow['summary']['unchanged']} unchanged"),
    ("  shrinkage applied", f"{stow['summary']['shrinkage_applied']} SKUs "
                            f"({stow['summary']['shrinkage_capped']} capped + flagged)"),
    ("  clumped demand", f"{stow['summary']['low_mover_poisson']} SKUs priced by negative binomial / burst floor"),
    ("HARRY GATOS", f"{hg['summary']['n_skus']} SKUs — {hg['summary']['increase']} up, "
                    f"{hg['summary']['decrease']} down, {hg['summary']['unchanged']} unchanged"),
    ("", ""),
    ("How to read it", "Sorted by biggest change first. Green = model wants more, orange = less."),
    ("Delta / Change %", "Recommended minus live par, as computed by the model on the date above."),
    ("Override (yellow)", "Hard-protected: drum reserves (additive), holds and zeros. The model can raise these but never lower them."),
    ("Burst floor", "Minimum to serve a realistic round: the 90th-percentile week in which the product actually sold."),
    ("Shrink /wk", "Units per week vanishing beyond recorded sales, measured from 12 Lightspeed stock counts (loss side only)."),
    ("Seasonal", "Week-of-year index — >1 means this week of the year runs hotter than the SKU's average."),
    ("Shared stock", "HG and Marilyna's sales of Stowaway stock are attributed to the Stowaway par. HG holds no par for [HG] SKUs."),
    ("", ""),
    ("Not yet applied", "Nothing here is uploaded to Lightspeed. Applying pars is a manual, approved step."),
    ("Bookings", f"Shadow only — {stow['bookings'].get('status', 'unavailable')}. Not added to any recommendation."),
    ("Christmas 2026", "Wed 23 Dec delivery must last to Wed 6 Jan = 21 day-units = 2.10x a normal order. See christmas_2026.md."),
]
for a, b in rows:
    ws.append([a, b])
for i in range(1, ws.max_row + 1):
    ws.cell(row=i, column=1).font = Font(name=FONT, bold=True, size=10)
    ws.cell(row=i, column=2).font = Font(name=FONT, size=10)
    ws.cell(row=i, column=2).alignment = Alignment(wrap_text=True, vertical="top")
ws["A1"].font = Font(name=FONT, bold=True, size=14)

wb.save(OUT)
print("wrote", OUT)
print("stow rows:", stow["summary"]["n_skus"], "| hg rows:", hg["summary"]["n_skus"])
